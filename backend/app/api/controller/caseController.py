# encoding: UTF-8
import os
import json

from sqlalchemy import and_, or_
from flask import g

from .baseCrudController import BaseCrudController
from ..model.caseModel import CaseReview, CaseSnapshot, Module, TestCase
from ..model.projectModel import Project
from ..model.userModel import User
from ..service.caseService import CaseService
from logger import logger


class CaseController(BaseCrudController):
    def module_list(self):
        project_id = self._get(self.req_data, 'projectId')
        parent_id = self._get(self.req_data, 'parentId')
        filters = []
        
        if project_id:
            filters.append(Module.project_id == int(project_id))
        if parent_id not in (None, ''):
            filters.append(Module.parent_id == int(parent_id))
        
        parent_module = Module.__table__.alias('parent')
        query = self.session.query(Module, parent_module.c.name.label('parent_name')).\
                outerjoin(parent_module, Module.parent_id == parent_module.c.id).\
                filter(*filters)
        
        if hasattr(Module, 'is_delete'):
            query = query.filter(Module.is_delete == 0)
        if hasattr(Module, 'status'):
            query = query.filter(Module.status == 1)
        
        total = query.count()
        
        page_num = int(self._get(self.req_data, 'pageNo', default=1))
        page_size = int(self._get(self.req_data, 'pageSize', default=200))
        
        query = query.order_by(Module.id)
        items = query.offset((page_num - 1) * page_size).limit(page_size).all()
        
        result_list = []
        for module, parent_name in items:
            module_dict = self.serialize(module, ['is_delete'])
            module_dict['parent_name'] = parent_name or ''
            result_list.append(module_dict)
        
        return {'list': result_list, 'total': total}

    def module_create(self):
        project_id = self._get(self.req_data, 'projectId')
        name = self._get(self.req_data, 'name')
        if not project_id or not name:
            return 0, 'projectId、name 为必传参数'
        add_info = {'project_id': project_id, 'parent_id': int(self._get(self.req_data, 'parentId', default=0)), 'name': name, 'sort_order': int(self._get(self.req_data, 'sortOrder', default=0)), 'path': self._get(self.req_data, 'path'), 'is_delete': 0}
        return CaseService.create(self.session, Module, add_info)

    def module_update(self):
        module_id = self._get(self.req_data, 'moduleId', 'id')
        if not module_id:
            return 0, 'moduleId 为必传参数'
        update_info = {}
        for req_key, column_key in [('parentId', 'parent_id'), ('name', 'name'), ('sortOrder', 'sort_order'), ('path', 'path')]:
            value = self._get(self.req_data, req_key)
            if value is not None:
                update_info[column_key] = value
        return CaseService.update_by_id(self.session, Module, module_id, update_info)

    def module_delete(self):
        module_id = self._get(self.req_data, 'moduleId', 'id')
        if not module_id:
            return 0, 'moduleId 为必传参数'
        return CaseService.delete_by_id(self.session, Module, module_id)

    def case_list(self):
        """分页查询用例列表，支持项目名称、用例标题、优先级、类型、状态、是否自动化、标签过滤。"""
        filters = []
        
        project_name = self._get(self.req_data, 'projectName')
        if project_name:
            filters.append(Project.name.like('%{}%'.format(project_name)))
        
        project_id = self._get(self.req_data, 'projectId')
        if project_id:
            filters.append(TestCase.project_id == int(project_id))
        
        module_name = self._get(self.req_data, 'moduleName', 'module_name')
        if module_name:
            filters.append(Module.name.like('%{}%'.format(module_name)))
        
        for req_key, column in [('moduleId', TestCase.module_id), ('priority', TestCase.priority), 
                                ('caseType', TestCase.case_type), ('isAuto', TestCase.is_auto)]:
            value = self._get(self.req_data, req_key)
            if value not in (None, ''):
                filters.append(column == int(value))
        is_ai_generated = self._get(self.req_data, 'isAiGenerated', 'is_ai_generated')
        is_ai_generated_val = None
        if is_ai_generated not in (None, ''):
            is_ai_generated_val = int(is_ai_generated)
            filters.append(TestCase.is_ai_generated == is_ai_generated_val)
        status = self._get(self.req_data, 'status')
        if status not in (None, ''):
            filters.append(TestCase.status == int(status))
        else:
            filters.append(TestCase.status != 0)
        
        keyword = self._get(self.req_data, 'keyword')
        if keyword:
            filters.append(TestCase.title.like('%{}%'.format(keyword)))
        
        tag = self._get(self.req_data, 'tag')
        if tag:
            filters.append(TestCase.tags.any(tag))
        
        created_by_name = self._get(self.req_data, 'createdBy')
        if created_by_name:
            filters.append(User.real_name.like('%{}%'.format(created_by_name)))
        
        query = self.session.query(TestCase, Project.name.label('project_name'), Module.name.label('module_name'), Module.path.label('module_path'), Module.status.label('module_status'), User.real_name.label('created_by_name')).\
                join(Project, TestCase.project_id == Project.id, isouter=True).\
                join(Module, TestCase.module_id == Module.id, isouter=True).\
                join(User, TestCase.created_by == User.id, isouter=True).\
                filter(*filters)
        
        if hasattr(TestCase, 'is_delete'):
            query = query.filter(TestCase.is_delete == 0)
        if hasattr(Project, 'is_delete'):
            query = query.filter(Project.is_delete == 0)
        if hasattr(Module, 'is_delete'):
            query = query.filter(or_(Module.is_delete == 0, Module.is_delete.is_(None)))
        if hasattr(Module, 'status'):
            module_status = self._get(self.req_data, 'moduleStatus', 'module_status')
            # AI 用例列表需要展示对应模块数据，不能因为模块状态变化把用例隐藏掉
            if is_ai_generated_val == 1:
                pass
            elif module_status not in (None, ''):
                # 非 AI 场景下，仍支持按模块状态过滤
                query = query.filter(Module.status == int(module_status))
            else:
                # 默认只查询状态为1的模块
                query = query.filter(or_(Module.status == 1, Module.status.is_(None)))
        if hasattr(User, 'is_delete'):
            query = query.filter(or_(User.is_delete == 0, User.is_delete.is_(None)))
        
        total = query.count()
        
        page_num = int(self._get(self.req_data, 'pageNo', 'page', default=1))
        page_size = int(self._get(self.req_data, 'pageSize', 'size', default=20))
        
        query = query.order_by(TestCase.id.desc())
        items = query.offset((page_num - 1) * page_size).limit(page_size).all()
        
        result_list = []
        for case, project_name, module_name, module_path, module_status, created_by_name in items:
            case_dict = self.serialize(case, ['is_delete'])
            case_dict['project_name'] = project_name or ''
            case_dict['module_name'] = module_name or ''
            case_dict['module_path'] = module_path or ''
            case_dict['module_id'] = case.module_id
            case_dict['module_status'] = module_status
            case_dict['case_key'] = case_dict.get('case_key', '')
            case_dict['created_by_name'] = created_by_name or ''
            if not case_dict.get('steps'):
                case_dict['steps'] = ''
            result_list.append(case_dict)
        
        return {'list': result_list, 'total': total}

    def case_detail(self):
        case_id = self._get(self.req_data, 'caseId', 'id')
        if not case_id:
            return {}, 'caseId 为必传参数'
        item = CaseService.get_by_id(self.session, TestCase, case_id)
        if not item:
            return {}, '未查询到对应用例！'
        result = self.serialize(item, ['is_delete'])
        if not result.get('steps'):
            result['steps'] = ''
        if item.module_id:
            module = self.session.query(Module).filter(Module.id == item.module_id).first()
            result['module_name'] = module.name if module else ''
        else:
            result['module_name'] = ''
        return result, ''

    def case_create(self):
        project_id = self._get(self.req_data, 'projectId')
        title = self._get(self.req_data, 'title')
        if not project_id or not title:
            return 0, 'projectId、title 为必传参数'
        steps_value = self._get(self.req_data, 'steps', default='')
        if isinstance(steps_value, (list, dict)):
            steps_value = ''
        product_id = self._get(self.req_data, 'productId')
        module_id = self._get(self.req_data, 'moduleId')
        add_info = {
            'project_id': project_id,
            'module_id': module_id,
            'case_key': self._get(self.req_data, 'caseKey') or CaseService.next_case_key(self.session, project_id, module_id, product_id),
            'title': title,
            'preconditions': self._get(self.req_data, 'preconditions'),
            'steps': steps_value,
            'expected_results': self._get(self.req_data, 'expectedResults'),
            'priority': int(self._get(self.req_data, 'priority', default=2)),
            'case_type': int(self._get(self.req_data, 'caseType', default=1)),
            'tags': self._get(self.req_data, 'tags', default=[]),
            'status': int(self._get(self.req_data, 'status', default=1)),
            'is_auto': int(self._get(self.req_data, 'isAuto', default=0)),
            'is_ai_generated': int(self._get(self.req_data, 'isAiGenerated', default=0)),
            'created_by': getattr(g, 'current_user_id', None),
            'is_delete': 0
        }
        return CaseService.create(self.session, TestCase, add_info)

    def case_update(self):
        """更新用例内容，只更新请求中传入的字段。"""
        case_id = self._get(self.req_data, 'caseId', 'id')
        if not case_id:
            return 0, 'caseId 为必传参数'
        update_info = {}
        mapping = [('moduleId', 'module_id'), ('caseKey', 'case_key'), ('title', 'title'), ('preconditions', 'preconditions'), ('expectedResults', 'expected_results'), ('priority', 'priority'), ('caseType', 'case_type'), ('tags', 'tags'), ('status', 'status'), ('isAuto', 'is_auto'), ('isAiGenerated', 'is_ai_generated')]
        for req_key, column_key in mapping:
            value = self._get(self.req_data, req_key)
            if value is not None:
                update_info[column_key] = value
        steps_value = self._get(self.req_data, 'steps')
        if steps_value is not None:
            if isinstance(steps_value, (list, dict)):
                steps_value = ''
            update_info['steps'] = steps_value
        return CaseService.update_by_id(self.session, TestCase, case_id, update_info)

    def case_delete(self):
        case_ids = self._get(self.req_data, 'caseIds', 'caseId', 'ids', 'id')
        if not case_ids:
            return {}, 'caseIds 为必传参数'
        if isinstance(case_ids, str):
            case_ids = [case_id.strip() for case_id in case_ids.split(',') if case_id.strip()]
        elif not isinstance(case_ids, list):
            case_ids = [case_ids]
        try:
            case_ids = list({int(case_id) for case_id in case_ids if str(case_id).strip()})
        except ValueError:
            return {}, 'caseIds 必须为数字数组或英文逗号分隔的数字字符串'
        if not case_ids:
            return {}, 'caseIds 为必传参数'
        delete_count = self.session.query(TestCase).filter(
            TestCase.id.in_(case_ids),
            TestCase.is_delete == 0
        ).update({'is_delete': 1}, synchronize_session=False)
        err = self.session.done(close=False)
        if err:
            return {}, f'删除失败！{err}'
        return {'caseIds': case_ids, 'deletedCount': delete_count}, ''

    def case_restore(self):
        case_ids = self._get(self.req_data, 'caseIds', 'caseId', 'ids', 'id')
        if not case_ids:
            return {}, 'caseIds 为必传参数'
        if isinstance(case_ids, str):
            case_ids = [case_id.strip() for case_id in case_ids.split(',') if case_id.strip()]
        elif not isinstance(case_ids, list):
            case_ids = [case_ids]
        try:
            case_ids = list({int(case_id) for case_id in case_ids if str(case_id).strip()})
        except ValueError:
            return {}, 'caseIds 必须为数字数组或英文逗号分隔的数字字符串'
        if not case_ids:
            return {}, 'caseIds 为必传参数'
        # 先获取需要恢复的用例对应的模块ID
        cases = self.session.query(TestCase).filter(
            TestCase.id.in_(case_ids),
            TestCase.status == 0,
            TestCase.is_delete == 0
        ).all()
        module_ids = list({case.module_id for case in cases if case.module_id})
        # 更新用例状态
        update_count = self.session.query(TestCase).filter(
            TestCase.id.in_(case_ids),
            TestCase.status == 0,
            TestCase.is_delete == 0
        ).update({'status': 1}, synchronize_session=False)
        # 更新模块状态为1，同时更新父模块状态
        if module_ids:
            # 获取所有需要更新的模块ID（包括父模块）
            all_module_ids = set(module_ids)
            current_ids = module_ids.copy()
            
            # 递归获取所有父模块ID
            while current_ids:
                parents = self.session.query(Module.parent_id).filter(
                    Module.id.in_(current_ids),
                    Module.parent_id != 0,
                    Module.is_delete == 0
                ).all()
                parent_ids = [p[0] for p in parents if p[0] not in all_module_ids]
                if not parent_ids:
                    break
                all_module_ids.update(parent_ids)
                current_ids = parent_ids
            
            # 更新所有模块（包括父模块）的状态为1
            self.session.query(Module).filter(
                Module.id.in_(list(all_module_ids)),
                Module.is_delete == 0
            ).update({'status': 1}, synchronize_session=False)
        err = self.session.done(close=False)
        if err:
            return {}, f'恢复失败！{err}'
        return {'caseIds': case_ids, 'updatedCount': update_count}, ''

    def snapshot_create(self):
        case_id = self._get(self.req_data, 'caseId')
        if not case_id:
            return 0, 'caseId 为必传参数'
        case_obj = CaseService.get_by_id(self.session, TestCase, case_id)
        if not case_obj:
            return 0, '未查询到对应用例！'
        version = CaseService.next_snapshot_version(self.session, case_id)
        snapshot = self.serialize(case_obj, ['is_delete'])
        if not snapshot.get('steps'):
            snapshot['steps'] = ''
        add_info = {'case_id': case_id, 'version': version, 'snapshot': snapshot, 'created_by': self._get(self.req_data, 'createdBy')}
        return CaseService.create(self.session, CaseSnapshot, add_info)

    def snapshot_list(self):
        """查询指定用例的快照历史。"""
        case_id = self._get(self.req_data, 'caseId')
        filters = [CaseSnapshot.case_id == int(case_id)] if case_id else []
        items, total = CaseService.list_by_filters(self.session, CaseSnapshot, filters, self._get(self.req_data, 'pageNo', default=1), self._get(self.req_data, 'pageSize', default=20), CaseSnapshot.created_time)
        return {'list': self.serialize_list(items), 'total': total}

    def review_create(self):
        case_id = self._get(self.req_data, 'caseId')
        reviewer_id = self._get(self.req_data, 'reviewerId')
        if not case_id or not reviewer_id:
            return 0, 'caseId、reviewerId 为必传参数'
        return CaseService.create(self.session, CaseReview, {'case_id': case_id, 'reviewer_id': reviewer_id, 'comments': self._get(self.req_data, 'comments')})

    def review_update(self):
        review_id = self._get(self.req_data, 'reviewId', 'id')
        if not review_id:
            return 0, 'reviewId 为必传参数'
        update_info = {}
        for req_key, column_key in [('status', 'status'), ('comments', 'comments'), ('diffContent', 'diff_content'), ('reviewedTime', 'reviewed_time')]:
            value = self._get(self.req_data, req_key)
            if value is not None:
                update_info[column_key] = value
        return CaseService.update_by_id(self.session, CaseReview, review_id, update_info, soft_delete=False)

    def review_list(self):
        """查询用例评审记录列表。"""
        case_id = self._get(self.req_data, 'caseId')
        filters = [CaseReview.case_id == int(case_id)] if case_id else []
        items, total = CaseService.list_by_filters(self.session, CaseReview, filters, self._get(self.req_data, 'pageNo', default=1), self._get(self.req_data, 'pageSize', default=20), CaseReview.created_time)
        return {'list': self.serialize_list(items), 'total': total}

    def case_import(self, file_path, project_id):
        """批量导入用例"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return 0, '请先安装 openpyxl 依赖'
        
        if not os.path.exists(file_path):
            return 0, '文件不存在'
        
        if not project_id:
            return 0, 'projectId 为必传参数'
        
        wb = load_workbook(file_path)
        sheet = wb.active
        
        headers = {}
        for col in range(1, sheet.max_column + 1):
            header = str(sheet.cell(row=1, column=col).value).strip() if sheet.cell(row=1, column=col).value else ''
            if header:
                headers[header] = col
        
        required_columns = ['所属模块', '用例标题', '前置条件', '步骤', '预期', '关键词', '优先级', '用例类型']
        for col in required_columns:
            if col not in headers:
                return 0, f'缺少必要列: {col}'
        
        module_name_to_id = {}
        existing_modules = self.session.query(Module).filter(Module.project_id == int(project_id), Module.is_delete == 0).all()
        for module in existing_modules:
            module_name_to_id[module.name] = module.id
        
        success_count = 0
        fail_count = 0
        fail_messages = []
        
        for row in range(2, sheet.max_row + 1):
            try:
                module_path_str = str(sheet.cell(row=row, column=headers['所属模块']).value).strip() if sheet.cell(row=row, column=headers['所属模块']).value else ''
                
                if not module_path_str:
                    fail_count += 1
                    fail_messages.append(f'第{row}行：所属模块为空')
                    continue
                
                module_names = [m.strip() for m in module_path_str.split('/') if m.strip()]
                
                if not module_names:
                    fail_count += 1
                    fail_messages.append(f'第{row}行：所属模块格式不正确')
                    continue
                
                parent_id = 0
                module_id = None
                
                for idx, module_name in enumerate(module_names):
                    if module_name in module_name_to_id:
                        parent_id = module_name_to_id[module_name]
                    else:
                        path_parts = module_names[:idx+1]
                        module_path = '/' + '/'.join(path_parts)
                        new_module = Module(
                            project_id=int(project_id),
                            parent_id=parent_id,
                            name=module_name,
                            path=module_path,
                            is_delete=0
                        )
                        self.session.add(new_module)
                        self.session.flush()
                        module_name_to_id[module_name] = new_module.id
                        parent_id = new_module.id
                
                module_id = parent_id
                
                title = str(sheet.cell(row=row, column=headers['用例标题']).value).strip() if sheet.cell(row=row, column=headers['用例标题']).value else ''
                preconditions = str(sheet.cell(row=row, column=headers['前置条件']).value).strip() if sheet.cell(row=row, column=headers['前置条件']).value else ''
                steps = str(sheet.cell(row=row, column=headers['步骤']).value).strip() if sheet.cell(row=row, column=headers['步骤']).value else ''
                expected_results = str(sheet.cell(row=row, column=headers['预期']).value).strip() if sheet.cell(row=row, column=headers['预期']).value else ''
                keywords = str(sheet.cell(row=row, column=headers['关键词']).value).strip() if sheet.cell(row=row, column=headers['关键词']).value else ''
                priority_str = str(sheet.cell(row=row, column=headers['优先级']).value).strip() if sheet.cell(row=row, column=headers['优先级']).value else '2'
                case_type_str = str(sheet.cell(row=row, column=headers['用例类型']).value).strip() if sheet.cell(row=row, column=headers['用例类型']).value else '1'
                
                if not title:
                    fail_count += 1
                    fail_messages.append(f'第{row}行：用例标题为空')
                    continue
                
                priority_map = {'P0': 0, 'P1': 1, 'P2': 2, 'P3': 3}
                priority = priority_map.get(priority_str, int(priority_str) if priority_str.isdigit() else 2)
                
                case_type_map = {'功能': 1, '性能': 2, '安全': 3, '接口': 4}
                case_type = case_type_map.get(case_type_str, int(case_type_str) if case_type_str.isdigit() else 1)
                
                tags = [k.strip() for k in keywords.split(',')] if keywords else []
                
                retry_count = 0
                max_retries = 5
                case_key = CaseService.next_case_key(self.session, project_id, module_id)
                
                while retry_count < max_retries:
                    try:
                        case = TestCase(
                            project_id=int(project_id),
                            module_id=module_id,
                            case_key=case_key,
                            title=title,
                            preconditions=preconditions,
                            steps=steps,
                            expected_results=expected_results,
                            priority=priority,
                            case_type=case_type,
                            tags=tags,
                            status=1,
                            is_auto=0,
                            is_ai_generated=0,
                            created_by=getattr(g, 'current_user_id', None),
                            is_delete=0
                        )
                        self.session.add(case)
                        self.session.flush()
                        success_count += 1
                        break
                    except Exception as e:
                        if 'duplicate key' in str(e).lower() or 'already exists' in str(e).lower():
                            logger.warning(f'case_import case_key冲突，重新生成：{case_key}, 错误：{str(e)}')
                            case_key = CaseService.next_case_key(self.session, project_id)
                            retry_count += 1
                        else:
                            raise
                
                if retry_count >= max_retries:
                    fail_count += 1
                    fail_messages.append(f'第{row}行：用例标题[{title}]导入失败：case_key生成失败')
                    
            except Exception as e:
                fail_count += 1
                fail_messages.append(f'第{row}行：导入失败 - {str(e)}')
        
        try:
            self.session.commit()
            msg = f'导入完成：成功{success_count}条，失败{fail_count}条'
            if fail_messages:
                msg += f'。失败详情：{"; ".join(fail_messages[:10])}'
                if len(fail_messages) > 10:
                    msg += f'...（共{len(fail_messages)}条）'
            return success_count, msg
        except Exception as e:
            self.session.rollback()
            return 0, f'提交失败：{str(e)}'

    @staticmethod
    def get_template_path():
        """获取模板文件路径"""
        return os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))), 'attachment', '用例导入模版.xlsx')
